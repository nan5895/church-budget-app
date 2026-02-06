"""
Church Worship Team Budget Manager
Zero-cost serverless web app using Streamlit + Google Sheets + Google Drive + Vision OCR
"""

import io
import re
import json
import datetime
from typing import Optional

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ──────────────────────────────────────────────
# Page Config
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="찬양팀 예산 관리",
    page_icon="⛪",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# Custom CSS — Sophisticated dark theme
# ──────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

:root {
    --primary: #6C63FF;
    --primary-light: #8B83FF;
    --accent: #00D2FF;
    --success: #00E676;
    --warning: #FFD600;
    --danger: #FF5252;
    --bg-dark: #0E1117;
    --bg-card: #1A1D26;
    --bg-card-hover: #22262F;
    --text-primary: #FAFAFA;
    --text-secondary: #A0AEC0;
    --border: #2D3748;
}

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Header gradient */
.main-header {
    background: linear-gradient(135deg, #6C63FF 0%, #00D2FF 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 2.4rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
    letter-spacing: -0.02em;
}

.sub-header {
    color: var(--text-secondary);
    font-size: 1rem;
    font-weight: 300;
    margin-bottom: 2rem;
}

/* Metric cards */
.metric-card {
    background: linear-gradient(145deg, #1A1D26 0%, #22262F 100%);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 1.5rem;
    text-align: center;
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}
.metric-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(108, 99, 255, 0.15);
}
.metric-label {
    color: var(--text-secondary);
    font-size: 0.85rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 0.5rem;
}
.metric-value {
    font-size: 1.8rem;
    font-weight: 700;
    color: var(--text-primary);
}
.metric-value.primary { color: var(--primary-light); }
.metric-value.success { color: var(--success); }
.metric-value.warning { color: var(--warning); }
.metric-value.danger  { color: var(--danger); }

/* Upload area */
.upload-zone {
    border: 2px dashed var(--primary);
    border-radius: 16px;
    padding: 2rem;
    text-align: center;
    background: rgba(108, 99, 255, 0.05);
    margin: 1rem 0;
}

/* Section divider */
.section-title {
    font-size: 1.3rem;
    font-weight: 600;
    background: linear-gradient(135deg, #6C63FF 0%, #00D2FF 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin: 2rem 0 1rem 0;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid var(--primary);
    display: inline-block;
}

/* Sidebar styling */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0E1117 0%, #1A1D26 100%);
}

/* Table styling */
.dataframe {
    border-radius: 12px !important;
    overflow: hidden;
}

/* Button styling */
.stButton > button {
    background: linear-gradient(135deg, #6C63FF 0%, #8B83FF 100%);
    color: white;
    border: none;
    border-radius: 10px;
    padding: 0.6rem 1.5rem;
    font-weight: 600;
    transition: all 0.2s ease;
}
.stButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 15px rgba(108, 99, 255, 0.4);
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px;
    padding: 8px 20px;
    font-weight: 500;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# Google API Connections (cached)
# ──────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/cloud-platform",
]


@st.cache_resource(ttl=300)
def get_credentials():
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return creds


@st.cache_resource(ttl=300)
def get_gspread_client():
    return gspread.authorize(get_credentials())


@st.cache_resource(ttl=300)
def get_drive_service():
    return build("drive", "v3", credentials=get_credentials())


@st.cache_resource(ttl=300)
def get_vision_client():
    from google.cloud import vision
    creds = get_credentials()
    return vision.ImageAnnotatorClient(credentials=creds)


def get_sheet():
    client = get_gspread_client()
    sheet_id = st.secrets["app_settings"]["spreadsheet_id"]
    spreadsheet = client.open_by_key(sheet_id)
    try:
        ws = spreadsheet.worksheet("Transactions")
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="Transactions", rows=1000, cols=10)
        ws.append_row([
            "Date", "Category", "Description", "Amount",
            "Payment Method", "Receipt URL", "OCR Amount", "Submitted By", "Timestamp"
        ])
    return ws


def get_budget_sheet():
    client = get_gspread_client()
    sheet_id = st.secrets["app_settings"]["spreadsheet_id"]
    spreadsheet = client.open_by_key(sheet_id)
    try:
        ws = spreadsheet.worksheet("Budget")
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="Budget", rows=100, cols=6)
        ws.append_row(["Category", "Monthly Budget", "Year", "Month", "Notes"])
    return ws


# ──────────────────────────────────────────────
# OCR via Google Cloud Vision
# ──────────────────────────────────────────────
def extract_amount_from_image(image_bytes: bytes) -> tuple[Optional[str], Optional[str]]:
    """Use Google Vision OCR to extract total amount from Korean receipt.
    Returns (amount_string, full_ocr_text) for debugging.
    """
    from google.cloud import vision

    client = get_vision_client()
    image = vision.Image(content=image_bytes)

    # document_text_detection is much better for structured docs like receipts
    response = client.document_text_detection(
        image=image,
        image_context={"language_hints": ["ko", "en"]},
    )

    if response.error.message:
        return None, None

    full_text = response.full_text_annotation.text if response.full_text_annotation else None
    if not full_text:
        return None, None

    lines = full_text.split("\n")
    keyword_candidates = []

    # Pattern: Korean won amounts like "27,190원" or "27,190" (with commas, reasonable length)
    # Max 10 digits to avoid picking up order numbers, phone numbers, etc.
    won_pattern = re.compile(r'([\d,]{1,12})(?:원)?')

    # Korean receipt keywords — must end with "금액" or be exact total keywords
    # to avoid matching "승인번호", "결제일시" etc.
    total_keywords = [
        "합계금액", "합계 금액", "총합계", "총 합계",
        "결제금액", "결제 금액", "승인금액", "승인 금액",
        "합계", "합 계",
        "총액", "총 액", "합산",
        "받을금액", "받을 금액", "청구금액", "청구 금액",
        "total", "amount",
    ]

    # Keywords that should NOT match (to filter "승인번호", "거래일시" etc.)
    exclude_keywords = ["번호", "일시", "종류", "개월", "상호", "주소", "등록", "상점", "정보"]

    def extract_won_amount(text: str) -> list[int]:
        """Extract reasonable Korean won amounts from text (100 ~ 99,999,999)."""
        results = []
        for m in won_pattern.finditer(text):
            raw = m.group(1).replace(",", "")
            if raw.isdigit():
                val = int(raw)
                if 100 <= val <= 99_999_999:
                    results.append(val)
        return results

    def line_has_exclude(text: str) -> bool:
        for ex in exclude_keywords:
            if ex in text:
                return True
        return False

    # Strategy 1: Find "합계금액" and look for amount with "원" suffix nearby
    # Handle receipts where labels and values are in separate blocks
    won_with_suffix = re.compile(r'([\d,]{1,12})원')
    for i, line in enumerate(lines):
        stripped = line.strip().replace(" ", "")
        if "합계금액" in stripped or "합계 금액" in stripped.replace(" ", ""):
            # Look for "XX,XXX원" pattern in same line
            m = won_with_suffix.search(line)
            if m:
                val = int(m.group(1).replace(",", ""))
                if 100 <= val <= 99_999_999:
                    return f"{val:,.0f}", full_text
            # Not on same line — scan ALL following lines for the first "원" suffixed amount
            for j in range(i + 1, len(lines)):
                m = won_with_suffix.search(lines[j])
                if m:
                    val = int(m.group(1).replace(",", ""))
                    if 100 <= val <= 99_999_999:
                        return f"{val:,.0f}", full_text

    # Strategy 2: General keyword search (same line or next line)
    for i, line in enumerate(lines):
        stripped = line.strip().replace(" ", "")
        if line_has_exclude(stripped):
            continue
        for priority, kw in enumerate(total_keywords):
            kw_clean = kw.replace(" ", "")
            if kw_clean in stripped:
                # Same line first
                amounts = extract_won_amount(line)
                if amounts:
                    for val in amounts:
                        keyword_candidates.append((priority, val))
                else:
                    # Next line
                    if i + 1 < len(lines) and not line_has_exclude(lines[i + 1]):
                        amounts = extract_won_amount(lines[i + 1])
                        for val in amounts:
                            keyword_candidates.append((priority, val))

    if keyword_candidates:
        keyword_candidates.sort(key=lambda x: (x[0], -x[1]))
        best_priority = keyword_candidates[0][0]
        best_group = [v for p, v in keyword_candidates if p == best_priority]
        return f"{max(best_group):,.0f}", full_text

    # Fallback: largest reasonable amount in entire text
    fallback = extract_won_amount(full_text)

    if fallback:
        return f"{max(fallback):,.0f}", full_text
    return None, full_text


# ──────────────────────────────────────────────
# Google Drive Upload
# ──────────────────────────────────────────────
def upload_to_drive(file_bytes: bytes, filename: str, mime_type: str) -> str:
    """Upload file to Google Drive shared folder, return public URL."""
    service = get_drive_service()
    folder_id = st.secrets["app_settings"]["drive_folder_id"]

    file_metadata = {
        "name": filename,
        "parents": [folder_id],
    }
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=True)

    # supportsAllDrives allows upload into folders shared with the service account
    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True,
    ).execute()

    # Make publicly viewable
    try:
        service.permissions().create(
            fileId=file["id"],
            body={"type": "anyone", "role": "reader"},
            supportsAllDrives=True,
        ).execute()
    except Exception:
        pass  # folder-level sharing may already cover this

    return file.get("webViewLink", f"https://drive.google.com/file/d/{file['id']}/view")


# ──────────────────────────────────────────────
# Data Loading
# ──────────────────────────────────────────────
@st.cache_data(ttl=60)
def load_transactions() -> pd.DataFrame:
    ws = get_sheet()
    data = ws.get_all_records()
    if not data:
        return pd.DataFrame(columns=[
            "Date", "Category", "Description", "Amount",
            "Payment Method", "Receipt URL", "OCR Amount", "Submitted By", "Timestamp"
        ])
    df = pd.DataFrame(data)
    if "Amount" in df.columns:
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df


@st.cache_data(ttl=60)
def load_budgets() -> pd.DataFrame:
    ws = get_budget_sheet()
    data = ws.get_all_records()
    if not data:
        return pd.DataFrame(columns=["Category", "Monthly Budget", "Year", "Month", "Notes"])
    df = pd.DataFrame(data)
    df["Monthly Budget"] = pd.to_numeric(df["Monthly Budget"], errors="coerce").fillna(0)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce").fillna(0).astype(int)
    # Handle legacy data without Month column
    if "Month" not in df.columns:
        df["Month"] = 0  # Default to annual budget (applies to all months)
    else:
        df["Month"] = pd.to_numeric(df["Month"], errors="coerce").fillna(0).astype(int)
    # Ensure Notes column exists
    if "Notes" not in df.columns:
        df["Notes"] = ""
    return df


def get_budget_for_month(budgets: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    """Get budget for a specific month. Only returns exact matches - no fallback."""
    if budgets.empty:
        return budgets

    # ONLY return budgets explicitly set for this specific year/month
    # No fallback to Month=0 or any other month
    monthly = budgets[(budgets["Year"] == year) & (budgets["Month"] == month)]

    return monthly


# ──────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────
def generate_excel_report(df: pd.DataFrame, budgets: pd.DataFrame = None) -> bytes:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = io.BytesIO()
    now = datetime.datetime.now()
    current_year = now.year
    current_month = now.month

    # Pre-calculate budget by category for this year
    budget_by_cat = {}
    if budgets is not None and not budgets.empty:
        for m in range(1, current_month + 1):
            month_budgets = get_budget_for_month(budgets, current_year, m)
            for _, row in month_budgets.iterrows():
                cat = row["Category"]
                budget_by_cat[cat] = budget_by_cat.get(cat, 0) + row["Monthly Budget"]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Sheet 1: Transactions (지출 내역) ──
        export_df = df.copy()
        if "Date" in export_df.columns:
            export_df["Date"] = export_df["Date"].astype(str)

        # Add Korean column names
        col_mapping = {
            "Date": "날짜",
            "Category": "카테고리",
            "Description": "설명",
            "Amount": "금액",
            "Payment Method": "결제수단",
            "Receipt URL": "영수증 URL",
            "OCR Amount": "OCR 인식금액",
            "Submitted By": "입력자",
            "Timestamp": "등록시간"
        }
        export_df = export_df.rename(columns=col_mapping)
        export_df.to_excel(writer, index=False, sheet_name="지출내역")

        # Add hyperlinks for receipt URLs
        ws = writer.sheets["지출내역"]
        url_col = None
        for idx, col in enumerate(export_df.columns, 1):
            if col == "영수증 URL":
                url_col = idx
                break
        if url_col:
            col_letter = get_column_letter(url_col)
            for row_idx in range(2, len(export_df) + 2):
                cell = ws[f"{col_letter}{row_idx}"]
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        # ── Sheet 2: Budget Settings (예산 설정) ──
        if budgets is not None and not budgets.empty:
            budget_df = budgets.copy()
            month_names = {0: "전체(연간)", 1: "1월", 2: "2월", 3: "3월", 4: "4월", 5: "5월",
                          6: "6월", 7: "7월", 8: "8월", 9: "9월", 10: "10월", 11: "11월", 12: "12월"}
            budget_df["Month"] = budget_df["Month"].apply(lambda x: month_names.get(x, str(x)))
            budget_df = budget_df.rename(columns={
                "Category": "카테고리",
                "Monthly Budget": "예산금액",
                "Year": "연도",
                "Month": "월",
                "Notes": "메모"
            })
            budget_df.to_excel(writer, index=False, sheet_name="예산설정")

        # ── Sheet 3: Category Summary (카테고리별 요약) ──
        if not df.empty and "Category" in df.columns and "Amount" in df.columns:
            summary = df.groupby("Category").agg({
                "Amount": ["sum", "count", "mean", "min", "max"]
            }).reset_index()
            summary.columns = ["카테고리", "총지출", "건수", "평균", "최소", "최대"]
            summary["총지출"] = summary["총지출"].apply(lambda x: round(x, 0))
            summary["평균"] = summary["평균"].apply(lambda x: round(x, 0))
            summary.to_excel(writer, index=False, sheet_name="카테고리별요약")

        # ── Sheet 4: Budget vs Actual (예산대비실적) ──
        if budget_by_cat and not df.empty:
            # This year's spending per category
            this_year_df = df[pd.to_datetime(df["Date"], errors="coerce").dt.year == current_year] if "Date" in df.columns else df
            cat_spent = this_year_df.groupby("Category")["Amount"].sum().reset_index() if not this_year_df.empty else pd.DataFrame(columns=["Category", "Amount"])

            comparison = pd.DataFrame(list(budget_by_cat.items()), columns=["카테고리", "누적예산"])
            comparison = comparison.merge(cat_spent.rename(columns={"Category": "카테고리", "Amount": "실제지출"}), on="카테고리", how="left").fillna(0)
            comparison["잔여예산"] = comparison["누적예산"] - comparison["실제지출"]
            comparison["집행률(%)"] = (comparison["실제지출"] / comparison["누적예산"] * 100).round(1)
            comparison["집행률(%)"] = comparison["집행률(%)"].replace([float('inf'), -float('inf')], 0).fillna(0)
            comparison["상태"] = comparison["잔여예산"].apply(lambda x: "✅ 정상" if x >= 0 else "⚠️ 초과")
            comparison.to_excel(writer, index=False, sheet_name="예산대비실적")

        # ── Sheet 5: Monthly Trend (월별추이) ──
        if not df.empty and "Date" in df.columns:
            df_copy = df.copy()
            df_copy["Date"] = pd.to_datetime(df_copy["Date"], errors="coerce")
            df_copy = df_copy.dropna(subset=["Date"])
            if not df_copy.empty:
                df_copy["연월"] = df_copy["Date"].dt.to_period("M").astype(str)
                monthly_trend = df_copy.groupby("연월").agg({
                    "Amount": ["sum", "count", "mean"]
                }).reset_index()
                monthly_trend.columns = ["연월", "총지출", "건수", "평균지출"]
                monthly_trend["총지출"] = monthly_trend["총지출"].apply(lambda x: round(x, 0))
                monthly_trend["평균지출"] = monthly_trend["평균지출"].apply(lambda x: round(x, 0))
                monthly_trend.to_excel(writer, index=False, sheet_name="월별추이")

        # ── Sheet 6: Payment Method Summary (결제수단별) ──
        if not df.empty and "Payment Method" in df.columns:
            payment_summary = df.groupby("Payment Method").agg({
                "Amount": ["sum", "count"]
            }).reset_index()
            payment_summary.columns = ["결제수단", "총금액", "건수"]
            payment_summary["비율(%)"] = (payment_summary["총금액"] / payment_summary["총금액"].sum() * 100).round(1)
            payment_summary.to_excel(writer, index=False, sheet_name="결제수단별")

        # ── Sheet 7: Report Summary (리포트요약) ──
        total_spent = df["Amount"].sum() if not df.empty else 0
        total_budget = sum(budget_by_cat.values()) if budget_by_cat else 0

        report_summary = pd.DataFrame({
            "항목": [
                "리포트 생성일",
                "총 거래 건수",
                "총 지출 금액",
                f"{current_year}년 누적 예산",
                "잔여 예산",
                "전체 집행률(%)"
            ],
            "값": [
                now.strftime("%Y-%m-%d %H:%M"),
                f"{len(df)}건",
                f"₩{total_spent:,.0f}",
                f"₩{total_budget:,.0f}",
                f"₩{total_budget - total_spent:,.0f}",
                f"{(total_spent / total_budget * 100):.1f}%" if total_budget > 0 else "N/A"
            ]
        })
        report_summary.to_excel(writer, index=False, sheet_name="리포트요약")

    return output.getvalue()


# ──────────────────────────────────────────────
# UI Helper
# ──────────────────────────────────────────────
def metric_card(label: str, value: str, color_class: str = "primary"):
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value {color_class}">{value}</div>
    </div>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p class="main-header" style="font-size:1.6rem;">⛪ 찬양팀 예산</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header" style="font-size:0.85rem;">Church Worship Team Budget</p>', unsafe_allow_html=True)
    st.divider()

    page = st.radio(
        "메뉴 / Menu",
        ["📊 대시보드", "📤 지출 입력", "📋 거래 내역", "⚙️ 예산 설정", "📥 리포트 다운로드"],
        label_visibility="collapsed",
    )
    st.divider()
    if st.button("🔄 데이터 새로고침"):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")
    st.caption("Powered by Streamlit · Google Sheets · Vision AI")

# ──────────────────────────────────────────────
# PAGE: Dashboard
# ──────────────────────────────────────────────
if page == "📊 대시보드":
    st.markdown('<p class="main-header">Budget Dashboard</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">찬양팀 예산 현황을 한눈에 확인하세요</p>', unsafe_allow_html=True)

    # Force fresh data load (no cache for dashboard)
    st.cache_data.clear()
    df = load_transactions()
    budgets = load_budgets()

    # Calculate metrics regardless of df being empty
    now = datetime.datetime.now()
    current_month = now.month
    current_year = now.year

    # Get budget for current month (from settings)
    current_month_budgets = get_budget_for_month(budgets, current_year, current_month)
    monthly_budget = current_month_budgets["Monthly Budget"].sum() if not current_month_budgets.empty else 0

    # Calculate previous months' budget (1월 ~ 지난달)
    prev_months_budget = 0
    for m in range(1, current_month):  # 1월부터 지난달까지
        month_budget = get_budget_for_month(budgets, current_year, m)
        prev_months_budget += month_budget["Monthly Budget"].sum() if not month_budget.empty else 0

    if not df.empty and "Date" in df.columns:
        # This month's spending
        this_month_mask = (df["Date"].dt.month == current_month) & (df["Date"].dt.year == current_year)
        this_month_spent = df.loc[this_month_mask, "Amount"].sum()

        # Previous months' spending this year (1월 ~ 지난달)
        prev_months_mask = (df["Date"].dt.month < current_month) & (df["Date"].dt.year == current_year)
        prev_months_spent = df.loc[prev_months_mask, "Amount"].sum()

        # This year's spending
        this_year_mask = df["Date"].dt.year == current_year
        total_spent_this_year = df.loc[this_year_mask, "Amount"].sum()

        tx_count = len(df)
    else:
        this_month_spent = 0
        prev_months_spent = 0
        total_spent_this_year = 0
        tx_count = 0

    # 이월 잔액 = 이전 달들의 예산 합계 - 이전 달들의 지출 합계
    carryover = prev_months_budget - prev_months_spent

    # 이번달 사용가능 예산 = 이번달 예산 + 이월 잔액
    available_budget = monthly_budget + carryover

    # 잔여 예산 = 이번달 사용가능 예산 - 이번달 지출
    remaining = available_budget - this_month_spent

    # KPI Cards — 6 metrics
    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card("이번달 예산", f"₩{monthly_budget:,.0f}", "primary")
    with c2:
        carryover_color = "success" if carryover >= 0 else "danger"
        metric_card("이월 잔액", f"₩{carryover:,.0f}", carryover_color)
    with c3:
        metric_card("사용가능 예산", f"₩{available_budget:,.0f}", "primary")

    c4, c5, c6 = st.columns(3)
    with c4:
        metric_card("이번달 지출", f"₩{this_month_spent:,.0f}", "warning")
    with c5:
        metric_card("잔여 예산", f"₩{remaining:,.0f}", "success" if remaining >= 0 else "danger")
    with c6:
        metric_card("거래 건수", f"{tx_count}건", "warning")

    # Show budget status message
    if monthly_budget == 0:
        st.warning(f"⚠️ {current_year}년 {current_month}월 예산이 설정되지 않았습니다. '예산 설정' 메뉴에서 예산을 설정해주세요.")

    if df.empty:
        st.info("아직 등록된 거래가 없습니다. '지출 입력' 메뉴에서 첫 거래를 등록하세요!")

    st.markdown("")

    # Charts row
    col_left, col_right = st.columns(2)

    # ── Chart 1: This Month's Budget vs Spending ──
    with col_left:
        st.markdown(f'<p class="section-title">{current_month}월 예산 대비 지출</p>', unsafe_allow_html=True)

        if monthly_budget > 0:
            # Calculate percentage
            spent_pct = min((this_month_spent / monthly_budget) * 100, 100) if monthly_budget > 0 else 0
            remaining_pct = max(100 - spent_pct, 0)
            over_amount = max(this_month_spent - monthly_budget, 0)

            # Donut chart for this month
            if over_amount > 0:
                # Over budget - show spent and over
                fig1 = go.Figure(data=[go.Pie(
                    labels=["지출 (예산 내)", "초과"],
                    values=[monthly_budget, over_amount],
                    hole=0.6,
                    marker_colors=["#FF5252", "#FF1744"],
                    textinfo="label+percent",
                    textfont=dict(color="#FAFAFA"),
                )])
                center_text = f"<b>초과</b><br>₩{over_amount:,.0f}"
            else:
                # Within budget - show spent and remaining
                fig1 = go.Figure(data=[go.Pie(
                    labels=["지출", "잔여"],
                    values=[this_month_spent, monthly_budget - this_month_spent],
                    hole=0.6,
                    marker_colors=["#FF5252", "#00E676"],
                    textinfo="label+percent",
                    textfont=dict(color="#FAFAFA"),
                )])
                center_text = f"<b>{spent_pct:.0f}%</b><br>사용"

            fig1.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#FAFAFA"),
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5),
                margin=dict(t=20, b=40, l=20, r=20),
                annotations=[dict(text=center_text, x=0.5, y=0.5, font_size=16, showarrow=False, font_color="#FAFAFA")]
            )
            st.plotly_chart(fig1, config={"displayModeBar": False}, use_container_width=True)
            st.caption(f"예산: ₩{monthly_budget:,.0f} | 지출: ₩{this_month_spent:,.0f} | 잔여: ₩{monthly_budget - this_month_spent:,.0f}")
        else:
            st.info(f"{current_month}월 예산이 설정되지 않았습니다.")
            if this_month_spent > 0:
                st.metric("이번달 지출", f"₩{this_month_spent:,.0f}")

    # ── Chart 2: Year-to-Date Total Budget vs Total Spending ──
    with col_right:
        st.markdown(f'<p class="section-title">{current_year}년 총 예산 대비 총 지출</p>', unsafe_allow_html=True)

        # Calculate year-to-date totals (only months 1 to current_month with budgets set)
        ytd_budget = 0
        ytd_spent = 0
        monthly_data = []

        for m in range(1, current_month + 1):
            m_budget = get_budget_for_month(budgets, current_year, m)
            m_budget_total = m_budget["Monthly Budget"].sum() if not m_budget.empty else 0

            if not df.empty:
                m_spent = df[(df["Date"].dt.month == m) & (df["Date"].dt.year == current_year)]["Amount"].sum()
            else:
                m_spent = 0

            if m_budget_total > 0:  # Only count months with budget set
                ytd_budget += m_budget_total
                ytd_spent += m_spent
                monthly_data.append({
                    "월": f"{m}월",
                    "예산": m_budget_total,
                    "지출": m_spent
                })

        if ytd_budget > 0:
            # Stacked bar chart showing budget vs spending by month
            if monthly_data:
                monthly_df = pd.DataFrame(monthly_data)

                fig2 = go.Figure()

                # Budget bars
                fig2.add_trace(go.Bar(
                    name="예산",
                    x=monthly_df["월"],
                    y=monthly_df["예산"],
                    marker_color="#6C63FF",
                    text=monthly_df["예산"].apply(lambda x: f"₩{x/10000:.0f}만"),
                    textposition="outside",
                    textfont=dict(color="#FAFAFA", size=10),
                ))

                # Spending bars
                fig2.add_trace(go.Bar(
                    name="지출",
                    x=monthly_df["월"],
                    y=monthly_df["지출"],
                    marker_color="#FF5252",
                    text=monthly_df["지출"].apply(lambda x: f"₩{x/10000:.0f}만"),
                    textposition="outside",
                    textfont=dict(color="#FAFAFA", size=10),
                ))

                fig2.update_layout(
                    barmode="group",
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#FAFAFA"),
                    xaxis=dict(gridcolor="#2D3748"),
                    yaxis=dict(gridcolor="#2D3748", title="원"),
                    margin=dict(t=30, b=20, l=20, r=20),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                )
                st.plotly_chart(fig2, config={"displayModeBar": False}, use_container_width=True)

            # Summary
            remaining_ytd = ytd_budget - ytd_spent
            usage_pct = (ytd_spent / ytd_budget * 100) if ytd_budget > 0 else 0
            st.caption(f"총 예산: ₩{ytd_budget:,.0f} | 총 지출: ₩{ytd_spent:,.0f} | 잔여: ₩{remaining_ytd:,.0f} ({usage_pct:.1f}% 사용)")
        else:
            st.info("설정된 예산이 없습니다.")


# ──────────────────────────────────────────────
# PAGE: Expense Input (with Receipt Upload + OCR)
# ──────────────────────────────────────────────
elif page == "📤 지출 입력":
    st.markdown('<p class="main-header">지출 입력</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">영수증을 업로드하면 금액이 자동 인식됩니다</p>', unsafe_allow_html=True)

    budgets = load_budgets()
    categories = budgets["Category"].tolist() if not budgets.empty else [
        "악기/장비", "음향장비", "악보/교재", "식비/간식", "교통비", "기타"
    ]

    # Receipt Upload Section
    st.markdown('<p class="section-title">📸 영수증 업로드 (선택)</p>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        "영수증 이미지를 드래그하거나 클릭하여 업로드",
        type=["jpg", "jpeg", "png", "webp"],
        help="JPG, PNG, WEBP 형식 지원",
    )

    ocr_amount = ""
    receipt_url = ""

    if uploaded_file is not None:
        col_img, col_ocr = st.columns([1, 1])

        with col_img:
            st.image(uploaded_file, caption="업로드된 영수증", width="stretch")

        with col_ocr:
            file_bytes = uploaded_file.getvalue()

            with st.spinner("🔍 OCR 분석 중..."):
                try:
                    detected, ocr_text = extract_amount_from_image(file_bytes)
                    if detected:
                        ocr_amount = detected
                        st.success(f"✅ 인식된 금액: **₩{detected}**")
                        st.caption("금액이 정확하지 않으면 아래에서 직접 수정하세요.")
                    else:
                        st.warning("금액을 자동 인식하지 못했습니다. 직접 입력해주세요.")
                    if ocr_text:
                        with st.expander("🔎 OCR 인식 원문 보기"):
                            st.text(ocr_text)
                except Exception as e:
                    st.warning(f"OCR 처리 중 오류: {e}")
                    st.caption("금액을 직접 입력해주세요.")

    # Input Form
    st.markdown('<p class="section-title">📝 거래 정보</p>', unsafe_allow_html=True)

    with st.form("expense_form", clear_on_submit=True):
        col1, col2 = st.columns(2)

        with col1:
            date = st.date_input("날짜", value=datetime.date.today())
            category = st.selectbox("카테고리", categories)
            description = st.text_input("설명", placeholder="예: 건반 수리비")

        with col2:
            default_amount = int(float(ocr_amount.replace(",", ""))) if ocr_amount else 0
            amount = st.number_input(
                "금액 (원)",
                min_value=0,
                value=default_amount,
                step=1000,
                help="OCR로 인식된 금액이 자동 입력됩니다",
            )
            payment_method = st.selectbox("결제 수단", ["카드", "현금", "계좌이체", "기타"])
            submitted_by = st.text_input("입력자", placeholder="이름")

        submitted = st.form_submit_button("💾 저장하기", width="stretch")

        if submitted:
            if amount <= 0:
                st.error("금액을 입력해주세요.")
            elif not description:
                st.error("설명을 입력해주세요.")
            else:
                with st.spinner("저장 중..."):
                    # Upload receipt if present
                    if uploaded_file is not None:
                        try:
                            file_bytes = uploaded_file.getvalue()
                            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            fname = f"receipt_{ts}_{uploaded_file.name}"
                            receipt_url = upload_to_drive(file_bytes, fname, uploaded_file.type)
                        except Exception as e:
                            st.warning(f"영수증 업로드 실패: {e}")
                            receipt_url = ""

                    # Append to Google Sheet
                    ws = get_sheet()
                    ws.append_row([
                        str(date),
                        category,
                        description,
                        amount,
                        payment_method,
                        receipt_url,
                        ocr_amount,
                        submitted_by,
                        datetime.datetime.now().isoformat(),
                    ])

                    st.cache_data.clear()
                    st.success("✅ 거래가 성공적으로 저장되었습니다!")
                    st.balloons()


# ──────────────────────────────────────────────
# PAGE: Transaction History
# ──────────────────────────────────────────────
elif page == "📋 거래 내역":
    st.markdown('<p class="main-header">거래 내역</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">전체 지출 기록을 조회 및 수정합니다</p>', unsafe_allow_html=True)

    df = load_transactions()
    budgets = load_budgets()

    # Get categories from budget settings
    budget_categories = budgets["Category"].tolist() if not budgets.empty else []
    default_categories = ["악기/장비", "음향장비", "악보/교재", "식비/간식", "교통비", "기타"]
    all_categories = list(set(budget_categories + default_categories))

    if df.empty:
        st.info("등록된 거래가 없습니다.")
    else:
        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            cats = ["전체"] + sorted(df["Category"].unique().tolist())
            sel_cat = st.selectbox("카테고리 필터", cats)
        with col_f2:
            if "Date" in df.columns and df["Date"].notna().any():
                min_date = df["Date"].min().date()
                max_date = df["Date"].max().date()
                date_range = st.date_input("기간", value=(min_date, max_date))
            else:
                date_range = None
        with col_f3:
            search = st.text_input("🔍 검색", placeholder="설명 검색...")

        filtered = df.copy()
        if sel_cat != "전체":
            filtered = filtered[filtered["Category"] == sel_cat]
        if date_range and len(date_range) == 2 and "Date" in filtered.columns:
            filtered = filtered[
                (filtered["Date"].dt.date >= date_range[0]) &
                (filtered["Date"].dt.date <= date_range[1])
            ]
        if search:
            filtered = filtered[
                filtered["Description"].astype(str).str.contains(search, case=False, na=False)
            ]

        st.markdown(f"**{len(filtered)}건** 조회됨 · 총 금액: **₩{filtered['Amount'].sum():,.0f}**")

        # Display table
        display_df = filtered.copy()
        if "Date" in display_df.columns:
            display_df["Date"] = display_df["Date"].dt.strftime("%Y-%m-%d")
        if "Amount" in display_df.columns:
            display_df["Amount"] = display_df["Amount"].apply(lambda x: f"₩{x:,.0f}")

        st.dataframe(
            display_df[["Date", "Category", "Description", "Amount", "Payment Method", "Submitted By"]],
            width="stretch",
            hide_index=True,
        )

        # ── Transaction Edit Section ──
        st.markdown('<p class="section-title">거래 수정</p>', unsafe_allow_html=True)
        st.info("💡 투명한 예산 관리를 위해 **금액은 수정할 수 없습니다**. 카테고리, 설명, 결제수단만 수정 가능합니다.")

        # Create a display string for each transaction
        df_with_idx = df.reset_index(drop=True)
        tx_options = []
        for i, row in df_with_idx.iterrows():
            date_str = row["Date"].strftime("%Y-%m-%d") if pd.notna(row["Date"]) else "날짜없음"
            amount_str = f"₩{row['Amount']:,.0f}"
            tx_options.append(f"{i}: [{date_str}] {row['Category']} - {row['Description'][:20]}... ({amount_str})")

        if tx_options:
            selected_tx = st.selectbox(
                "수정할 거래 선택",
                range(len(df_with_idx)),
                format_func=lambda i: tx_options[i],
                key="tx_edit_select"
            )

            sel_row = df_with_idx.iloc[selected_tx]

            with st.form("edit_transaction_form"):
                st.markdown(f"**선택된 거래 금액: ₩{sel_row['Amount']:,.0f}** (수정 불가)")

                etc1, etc2 = st.columns(2)
                with etc1:
                    # Category edit - combine existing categories with budget categories
                    existing_cats = df["Category"].unique().tolist()
                    combined_cats = sorted(list(set(existing_cats + all_categories)))
                    current_cat_idx = combined_cats.index(sel_row["Category"]) if sel_row["Category"] in combined_cats else 0
                    edit_category = st.selectbox("카테고리", combined_cats, index=current_cat_idx)
                    edit_description = st.text_input("설명", value=str(sel_row.get("Description", "")))
                with etc2:
                    payment_methods = ["카드", "현금", "계좌이체", "기타"]
                    current_pm = sel_row.get("Payment Method", "카드")
                    pm_idx = payment_methods.index(current_pm) if current_pm in payment_methods else 0
                    edit_payment = st.selectbox("결제수단", payment_methods, index=pm_idx)
                    edit_submitted_by = st.text_input("입력자", value=str(sel_row.get("Submitted By", "")))

                tfc1, tfc2 = st.columns(2)
                with tfc1:
                    if st.form_submit_button("💾 수정 저장", use_container_width=True):
                        ws = get_sheet()
                        sheet_row = selected_tx + 2  # +1 for header, +1 for 0-index
                        # Update only Category, Description, Payment Method, Submitted By (columns B, C, E, H)
                        ws.update(f"B{sheet_row}", [[edit_category]])
                        ws.update(f"C{sheet_row}", [[edit_description]])
                        ws.update(f"E{sheet_row}", [[edit_payment]])
                        ws.update(f"H{sheet_row}", [[edit_submitted_by]])
                        st.cache_data.clear()
                        st.success(f"✅ 거래가 수정되었습니다.")
                        st.rerun()
                with tfc2:
                    if st.form_submit_button("🗑️ 삭제", use_container_width=True):
                        ws = get_sheet()
                        sheet_row = selected_tx + 2
                        ws.delete_rows(sheet_row)
                        st.cache_data.clear()
                        st.success("✅ 거래가 삭제되었습니다.")
                        st.rerun()


# ──────────────────────────────────────────────
# PAGE: Budget Settings
# ──────────────────────────────────────────────
elif page == "⚙️ 예산 설정":
    st.markdown('<p class="main-header">예산 설정</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">카테고리별 월 예산을 설정합니다 (각 월별로 설정 필요)</p>', unsafe_allow_html=True)

    # Month 1-12 only (no Month=0 "annual" option - each month must be set explicitly)
    month_names = {1: "1월", 2: "2월", 3: "3월", 4: "4월", 5: "5월",
                   6: "6월", 7: "7월", 8: "8월", 9: "9월", 10: "10월", 11: "11월", 12: "12월"}
    # For display of old data with Month=0
    month_names_display = {0: "⚠️ 미지정", **month_names}

    now = datetime.datetime.now()
    current_year = now.year
    current_month = now.month

    # Load fresh data directly from Google Sheet (bypass cache)
    def load_budgets_fresh():
        ws = get_budget_sheet()
        data = ws.get_all_records()
        if not data:
            return pd.DataFrame(columns=["Category", "Monthly Budget", "Year", "Month", "Notes"])
        df = pd.DataFrame(data)
        df["Monthly Budget"] = pd.to_numeric(df["Monthly Budget"], errors="coerce").fillna(0)
        df["Year"] = pd.to_numeric(df["Year"], errors="coerce").fillna(0).astype(int)
        if "Month" not in df.columns:
            df["Month"] = 0
        else:
            df["Month"] = pd.to_numeric(df["Month"], errors="coerce").fillna(0).astype(int)
        if "Notes" not in df.columns:
            df["Notes"] = ""
        return df.reset_index(drop=True)

    budgets = load_budgets_fresh()

    # ── Data Migration: Fix Month=0 entries ──
    month0_entries = budgets[budgets["Month"] == 0]
    if not month0_entries.empty:
        st.error(f"⚠️ 월이 미지정(0)인 항목이 {len(month0_entries)}건 있습니다. 예산이 제대로 집계되지 않습니다.")
        if st.button("🔧 1월 예산으로 자동 변환", key="migrate_btn"):
            ws = get_budget_sheet()
            for idx, row in month0_entries.iterrows():
                sheet_row = idx + 2  # header row + 0-based index
                # Update Month column (D) to 1
                ws.update(f"D{sheet_row}", [[1]])
            st.cache_data.clear()
            st.success(f"✅ {len(month0_entries)}건의 예산이 1월로 변환되었습니다.")
            st.rerun()

    # ── Total Budget Summary ──
    st.markdown('<p class="section-title">예산 요약</p>', unsafe_allow_html=True)

    if not budgets.empty:
        # Calculate this month's budget
        this_month_budgets = budgets[(budgets["Year"] == current_year) & (budgets["Month"] == current_month)]
        this_month_total = this_month_budgets["Monthly Budget"].sum() if not this_month_budgets.empty else 0

        # Calculate yearly total (only explicitly set months)
        yearly_budgets = budgets[(budgets["Year"] == current_year) & (budgets["Month"] >= 1) & (budgets["Month"] <= 12)]
        yearly_total = yearly_budgets["Monthly Budget"].sum() if not yearly_budgets.empty else 0

        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            metric_card(f"{current_month}월 예산", f"₩{this_month_total:,.0f}", "primary")
        with sc2:
            metric_card(f"{current_year}년 총예산", f"₩{yearly_total:,.0f}", "primary")
        with sc3:
            cat_count = len(this_month_budgets["Category"].unique()) if not this_month_budgets.empty else 0
            metric_card("카테고리 수", f"{cat_count}개", "warning")

        # Category breakdown for this month
        if not this_month_budgets.empty:
            st.markdown(f"**{current_year}년 {current_month}월 카테고리별 예산:**")
            cat_summary = this_month_budgets.groupby("Category")["Monthly Budget"].sum().reset_index()
            cat_summary = cat_summary.rename(columns={"Category": "카테고리", "Monthly Budget": "예산"})
            cat_summary["예산"] = cat_summary["예산"].apply(lambda x: f"₩{x:,.0f}")
            st.dataframe(cat_summary, width="stretch", hide_index=True)
        else:
            st.warning(f"⚠️ {current_year}년 {current_month}월 예산이 설정되지 않았습니다. 아래에서 예산을 추가해주세요.")
    else:
        st.info("설정된 예산이 없습니다. 아래에서 예산을 추가해주세요.")

    st.markdown("---")

    # ── Current Budget List with Filters ──
    if not budgets.empty:
        st.markdown('<p class="section-title">현재 예산 목록</p>', unsafe_allow_html=True)

        # Filters
        filter_col1, filter_col2, filter_col3 = st.columns(3)
        with filter_col1:
            years_in_data = sorted(budgets["Year"].unique().tolist(), reverse=True)
            filter_year = st.selectbox("연도 필터", ["전체"] + years_in_data, key="budget_filter_year")
        with filter_col2:
            filter_month = st.selectbox("월 필터", ["전체"] + list(month_names.keys()), format_func=lambda x: "전체" if x == "전체" else month_names.get(x, str(x)), key="budget_filter_month")
        with filter_col3:
            cats_in_data = sorted(budgets["Category"].unique().tolist())
            filter_cat = st.selectbox("카테고리 필터", ["전체"] + cats_in_data, key="budget_filter_cat")

        # Apply filters for display only
        filtered_budgets = budgets.copy()
        if filter_year != "전체":
            filtered_budgets = filtered_budgets[filtered_budgets["Year"] == filter_year]
        if filter_month != "전체":
            filtered_budgets = filtered_budgets[filtered_budgets["Month"] == filter_month]
        if filter_cat != "전체":
            filtered_budgets = filtered_budgets[filtered_budgets["Category"] == filter_cat]

        if not filtered_budgets.empty:
            display_b = filtered_budgets.copy()
            display_b["Monthly Budget"] = display_b["Monthly Budget"].apply(lambda x: f"₩{x:,.0f}")
            display_b["Month"] = display_b["Month"].apply(lambda x: month_names_display.get(x, str(x)))
            display_b = display_b.rename(columns={"Monthly Budget": "예산", "Year": "연도", "Month": "월", "Category": "카테고리", "Notes": "메모"})
            st.dataframe(display_b[["카테고리", "예산", "연도", "월", "메모"]], width="stretch", hide_index=True)
            st.caption(f"총 {len(filtered_budgets)}건")

            # Warning for Month=0 entries
            old_entries = budgets[budgets["Month"] == 0]
            if not old_entries.empty:
                st.warning(f"⚠️ '미지정(월=0)' 항목이 {len(old_entries)}건 있습니다. 이 항목들은 예산에 반영되지 않습니다. 아래 '예산 수정'에서 올바른 월(1-12)로 수정해주세요.")
        else:
            st.info("필터 조건에 맞는 예산이 없습니다.")

        # ── Edit existing budget ──
        st.markdown('<p class="section-title">예산 수정</p>', unsafe_allow_html=True)

        # Create selection options with row index stored
        budget_options = []
        for idx, row in budgets.iterrows():
            label = f"{row['Category']} ({row['Year']}년 {month_names_display.get(row['Month'], '?')})"
            budget_options.append((idx, label, row))

        selected_option = st.selectbox(
            "수정할 항목 선택",
            range(len(budget_options)),
            format_func=lambda i: budget_options[i][1],
            key="edit_select",
        )

        edit_idx, _, sel = budget_options[selected_option]

        # Check if current entry has Month=0 (needs fixing)
        current_month_val = int(sel["Month"])
        if current_month_val == 0:
            st.error(f"⚠️ 이 항목은 월이 '미지정(0)'입니다. 아래에서 올바른 월(1-12)을 선택해주세요.")

        # Display current values
        st.markdown(f"**현재 값:** {sel['Category']} | ₩{sel['Monthly Budget']:,.0f} | {sel['Year']}년 {month_names_display.get(current_month_val, '?')}")

        with st.form("edit_budget_form", clear_on_submit=False):
            ec1, ec2 = st.columns(2)
            with ec1:
                edit_cat = st.text_input("카테고리", value=str(sel["Category"]), key="edit_cat")
                edit_budget = st.number_input("예산 (원)", value=int(sel["Monthly Budget"]), min_value=0, step=10000, key="edit_budget")
            with ec2:
                edit_year = st.number_input("연도", value=int(sel["Year"]), min_value=2020, max_value=2100, key="edit_year")
                # Only allow months 1-12
                month_options = list(month_names.keys())  # [1, 2, ..., 12]
                # If current month is 0, default to current month
                default_month = current_month_val if current_month_val in month_options else current_month
                edit_month = st.selectbox(
                    "월 (1-12월 선택)",
                    options=month_options,
                    format_func=lambda x: month_names[x],
                    index=month_options.index(default_month),
                    key="edit_month"
                )
            edit_notes = st.text_input("메모", value=str(sel.get("Notes", "") or ""), key="edit_notes")

            st.info("💡 각 월의 예산은 해당 월에만 적용됩니다. 여러 달에 같은 예산을 적용하려면 각 월별로 추가해주세요.")

            fc1, fc2 = st.columns(2)
            with fc1:
                if st.form_submit_button("💾 수정 저장", use_container_width=True):
                    ws = get_budget_sheet()
                    # Google Sheet row = DataFrame index + 2 (1 for header, 1 for 0-based index)
                    sheet_row = edit_idx + 2
                    ws.update(f"A{sheet_row}:E{sheet_row}", [[edit_cat, edit_budget, edit_year, edit_month, edit_notes]])
                    st.cache_data.clear()
                    st.success(f"✅ '{edit_cat}' 예산이 수정되었습니다. ({edit_year}년 {month_names[edit_month]})")
                    st.rerun()
            with fc2:
                if st.form_submit_button("🗑️ 삭제", use_container_width=True):
                    ws = get_budget_sheet()
                    sheet_row = edit_idx + 2
                    ws.delete_rows(sheet_row)
                    st.cache_data.clear()
                    st.success(f"✅ '{sel['Category']}' 예산이 삭제되었습니다.")
                    st.rerun()

    st.markdown('<p class="section-title">예산 추가</p>', unsafe_allow_html=True)

    with st.form("budget_form", clear_on_submit=True):
        bc1, bc2 = st.columns(2)
        with bc1:
            new_cat = st.text_input("카테고리명", placeholder="예: 악기/장비")
            new_budget = st.number_input("예산 (원)", min_value=0, step=10000, key="new_budget")
        with bc2:
            new_year = st.number_input("연도", value=datetime.date.today().year, min_value=2020, max_value=2100, key="new_year")
            add_month_options = list(month_names.keys())  # [1, 2, ..., 12]
            new_month = st.selectbox(
                "월 (1-12월 선택)",
                options=add_month_options,
                format_func=lambda x: month_names[x],
                index=add_month_options.index(current_month),  # default to current month
                key="new_month_select"
            )

        notes = st.text_input("메모", placeholder="선택사항", key="new_notes")

        st.info("💡 각 월의 예산은 해당 월에만 적용됩니다. 여러 달에 같은 예산이 필요하면 각 월별로 추가해주세요.")

        if st.form_submit_button("➕ 예산 추가", use_container_width=True):
            if new_cat and new_budget > 0:
                ws = get_budget_sheet()
                ws.append_row([new_cat, new_budget, new_year, new_month, notes])
                st.cache_data.clear()
                st.success(f"✅ '{new_cat}' 예산이 추가되었습니다 ({new_year}년 {month_names[new_month]})")
                st.rerun()
            else:
                st.error("카테고리명과 예산 금액을 입력해주세요.")


# ──────────────────────────────────────────────
# PAGE: Report Download
# ──────────────────────────────────────────────
elif page == "📥 리포트 다운로드":
    st.markdown('<p class="main-header">리포트 다운로드</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">예산 및 지출 현황을 포함한 상세 리포트를 다운로드합니다</p>', unsafe_allow_html=True)

    # Force fresh data
    st.cache_data.clear()
    df = load_transactions()
    budgets = load_budgets()

    now = datetime.datetime.now()
    current_year = now.year
    current_month = now.month

    # Report Overview
    st.markdown('<p class="section-title">리포트 개요</p>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        total_tx = len(df) if not df.empty else 0
        metric_card("총 거래 건수", f"{total_tx}건", "primary")

    with col2:
        total_spent = df["Amount"].sum() if not df.empty else 0
        metric_card("총 지출 금액", f"₩{total_spent:,.0f}", "warning")

    with col3:
        # Calculate cumulative budget
        cumulative_budget = 0
        for m in range(1, current_month + 1):
            month_budget = get_budget_for_month(budgets, current_year, m)
            cumulative_budget += month_budget["Monthly Budget"].sum() if not month_budget.empty else 0
        remaining = cumulative_budget - total_spent
        metric_card("잔여 예산", f"₩{remaining:,.0f}", "success" if remaining >= 0 else "danger")

    st.markdown("")

    # Category Summary
    if not df.empty and "Category" in df.columns:
        st.markdown('<p class="section-title">카테고리별 지출 요약</p>', unsafe_allow_html=True)
        summary = df.groupby("Category")["Amount"].agg(["sum", "count", "mean"]).reset_index()
        summary.columns = ["카테고리", "총액", "건수", "평균"]
        summary["총액"] = summary["총액"].apply(lambda x: f"₩{x:,.0f}")
        summary["평균"] = summary["평균"].apply(lambda x: f"₩{x:,.0f}")
        st.dataframe(summary, width="stretch", hide_index=True)

    # Budget vs Actual Preview
    if not budgets.empty:
        st.markdown('<p class="section-title">예산 대비 실적 요약</p>', unsafe_allow_html=True)

        budget_by_cat = {}
        for m in range(1, current_month + 1):
            month_budgets = get_budget_for_month(budgets, current_year, m)
            for _, row in month_budgets.iterrows():
                cat = row["Category"]
                budget_by_cat[cat] = budget_by_cat.get(cat, 0) + row["Monthly Budget"]

        if not df.empty:
            cat_spent = df.groupby("Category")["Amount"].sum().reset_index()
            comparison = pd.DataFrame(list(budget_by_cat.items()), columns=["카테고리", "누적예산"])
            comparison = comparison.merge(cat_spent.rename(columns={"Category": "카테고리", "Amount": "지출"}), on="카테고리", how="left").fillna(0)
            comparison["잔여"] = comparison["누적예산"] - comparison["지출"]
            comparison["집행률"] = (comparison["지출"] / comparison["누적예산"] * 100).round(1)
            comparison["집행률"] = comparison["집행률"].replace([float('inf'), -float('inf')], 0).fillna(0)
            comparison["집행률"] = comparison["집행률"].apply(lambda x: f"{x:.1f}%")
            comparison["누적예산"] = comparison["누적예산"].apply(lambda x: f"₩{x:,.0f}")
            comparison["지출"] = comparison["지출"].apply(lambda x: f"₩{x:,.0f}")
            comparison["잔여"] = comparison["잔여"].apply(lambda x: f"₩{x:,.0f}")
            st.dataframe(comparison, width="stretch", hide_index=True)

    # Report Contents Info
    st.markdown('<p class="section-title">리포트 포함 내용</p>', unsafe_allow_html=True)
    st.markdown("""
    다운로드되는 Excel 파일에는 다음 시트가 포함됩니다:

    1. **지출내역** - 전체 거래 목록 (날짜, 카테고리, 설명, 금액, 결제수단 등)
    2. **예산설정** - 설정된 예산 목록 (카테고리별 연간/월간 예산)
    3. **카테고리별요약** - 카테고리별 총지출, 건수, 평균, 최소, 최대
    4. **예산대비실적** - 예산 대비 실제 지출 현황 및 집행률
    5. **월별추이** - 월별 지출 추이
    6. **결제수단별** - 결제수단별 지출 현황
    7. **리포트요약** - 주요 지표 요약
    """)

    st.markdown("")

    if df.empty and budgets.empty:
        st.info("다운로드할 데이터가 없습니다. 먼저 예산 설정 또는 지출을 입력해주세요.")
    else:
        excel_data = generate_excel_report(df, budgets)
        today = datetime.date.today().strftime("%Y%m%d")

        st.download_button(
            label="📥 Excel 리포트 다운로드",
            data=excel_data,
            file_name=f"찬양팀_예산리포트_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.officedocument",
            width="stretch",
        )
